import pandas as pd
import io
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def gerar_excel(df: pd.DataFrame, template_bytes: bytes, header_row: int = 2) -> bytes:
    """
    Insere o DataFrame no template Excel preservando formatação.
    Retorna bytes prontos para download.
    """
    wb = load_workbook(io.BytesIO(template_bytes))
    ws = wb.active

    # Linha de início dos dados (após o cabeçalho)
    data_start_row = header_row

    # Limpar dados existentes (manter apenas cabeçalho)
    max_row = ws.max_row
    for row in ws.iter_rows(min_row=data_start_row, max_row=max_row):
        for cell in row:
            cell.value = None

    # Inserir dados do DataFrame
    for r_idx, row_data in enumerate(dataframe_to_rows(df, index=False, header=False)):
        for c_idx, valor in enumerate(row_data, start=1):
            ws.cell(row=data_start_row + r_idx, column=c_idx, value=valor)

    # Salvar em memória
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
